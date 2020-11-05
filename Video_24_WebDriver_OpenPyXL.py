import openpyxl

path = "C:/Users/Manoj/Desktop/Python - Selenium Practice/Drivers/test.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column


print(rows)
print(cols)

for i in range(1,rows+1):
    for j in range(1,cols+1):
        print(sheet.cell(row=i,column=j).value,end='  ')
    print()